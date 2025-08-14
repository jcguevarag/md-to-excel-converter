#!/usr/bin/env python3
"""
Markdown Table to Excel Converter

A utility script that converts the first Markdown table found in a file to a 
formatted Excel (.xlsx) spreadsheet. The script preserves basic Markdown 
formatting (bold, italic) and handles HTML line breaks within cells.

Dependencies:
    - pandas: For data manipulation and DataFrame operations
    - openpyxl: For Excel file creation and formatting (optional but recommended)

Usage:
    python md_to_excel.py -i input.md -o output.xlsx

Author: Juan C. Guevara G. (supported by AI)
Version: 1.0.0
"""

import argparse
import os
import re
from typing import List, Optional, Tuple

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def unescape_markdown_cell(text: str) -> str:
    """
    Unescape Markdown special characters in table cell content.
    
    Converts escaped pipes (\\|) back to regular pipes and escaped backslashes
    (\\\\) back to single backslashes. Order of operations matters to avoid
    double-unescaping.
    
    Args:
        text: The raw cell content containing escaped characters
        
    Returns:
        The unescaped cell content
    """
    return text.replace('\\|', '|').replace('\\\\', '\\')


def preprocess_cell_content(text: Optional[str]) -> str:
    """
    Preprocess cell content by converting HTML breaks to newlines.
    
    Replaces various forms of HTML <br> tags with newline characters while
    preserving other Markdown formatting markers like **, *, _.
    
    Args:
        text: The raw cell content, may be None
        
    Returns:
        Cleaned cell content with HTML breaks converted to newlines
    """
    if text is None:
        return ""
    
    text = str(text)
    # Replace <br> variants (case-insensitive) with newline
    text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
    return text.strip()


def parse_markdown_table_line(line: str) -> Optional[List[str]]:
    """
    Parse a single line of a Markdown table into its constituent cells.
    
    Identifies valid table rows by checking for proper pipe delimiters and
    splits the content while respecting escaped pipes. Each cell is then
    preprocessed to handle HTML content and Markdown escaping.
    
    Args:
        line: A single line from the Markdown file
        
    Returns:
        List of preprocessed cell contents, or None if the line is not a valid table row
    """
    line = line.strip()
    if not line.startswith('|') or not line.endswith('|'):
        return None
    
    inner = line[1:-1]
    
    # Split only on unescaped pipes: a '|' not preceded by a backslash
    raw_cells = re.split(r'(?<!\\)\|', inner)
    
    # Unescape special characters and preprocess each cell
    preprocessed_cells = [
        preprocess_cell_content(unescape_markdown_cell(cell.strip()))
        for cell in raw_cells
    ]
    
    return preprocessed_cells


def is_separator_line(line: str) -> bool:
    """
    Check if a line is a Markdown table separator row.
    
    Separator lines contain dashes and optional colons to indicate column
    alignment (e.g., |---|:---:|---:|). This function validates the format
    of such lines.
    
    Args:
        line: A single line from the Markdown file
        
    Returns:
        True if the line is a valid table separator, False otherwise
    """
    line = line.strip()
    if not line.startswith('|') or not line.endswith('|'):
        return False
    
    inner = line[1:-1]
    cols = [cell.strip() for cell in re.split(r'(?<!\\)\|', inner)]
    
    if not cols:
        return False
    
    # Validate each column separator (allows dashes, colons, and spaces)
    return all(re.fullmatch(r'[:\s-]*-+[:\s-]*', col or '-') for col in cols)


def apply_cell_formatting(cell, value: str) -> Tuple[str, Optional[Font]]:
    """
    Apply Markdown formatting to an Excel cell.
    
    Detects Markdown formatting markers (**, *, _) and returns the cleaned
    text along with the appropriate Excel font formatting.
    
    Args:
        cell: The Excel cell object to format
        value: The cell content as a string
        
    Returns:
        Tuple of (cleaned_text, font_object) where font_object may be None
    """
    cleaned_value = value
    font_to_apply = None
    
    # Check for bold formatting (**)
    if value.startswith('**') and value.endswith('**') and len(value) > 4:
        cleaned_value = value[2:-2].strip()
        font_to_apply = Font(bold=True)
    # Check for italic formatting (* or _)
    elif ((value.startswith('*') and value.endswith('*')) or 
          (value.startswith('_') and value.endswith('_'))) and len(value) > 2:
        cleaned_value = value[1:-1].strip()
        font_to_apply = Font(italic=True)
    
    return cleaned_value, font_to_apply


def parse_markdown_table(lines: List[str]) -> Tuple[List[str], List[List[str]], bool]:
    """
    Extract the first Markdown table from a list of file lines.
    
    Parses line by line looking for a valid Markdown table structure:
    header row, separator row, and data rows. Stops parsing when the
    table ends or an invalid structure is encountered.
    
    Args:
        lines: List of lines from the Markdown file
        
    Returns:
        Tuple of (headers, data_rows, success) where:
        - headers: List of column headers
        - data_rows: List of lists containing row data
        - success: Boolean indicating if parsing was successful
    """
    headers = []
    data_rows = []
    header_found = False
    separator_found = False
    table_found = False
    
    print("Processing Markdown content...")
    
    for i, line in enumerate(lines):
        # Check for separator first
        if header_found and not separator_found:
            if is_separator_line(line):
                separator_found = True
                print(f"  Found separator on line {i+1}")
                continue
        
        # Parse potential table row
        parsed_cols = parse_markdown_table_line(line)
        
        if parsed_cols is not None:
            if not header_found:
                headers = parsed_cols
                header_found = True
                table_found = True
                print(f"  Found header on line {i+1}: {headers}")
            elif separator_found:
                if len(parsed_cols) != len(headers):
                    print(f"  Warning: Row {i+1} has {len(parsed_cols)} columns, "
                          f"expected {len(headers)}")
                data_rows.append(parsed_cols)
            elif header_found and not separator_found:
                print(f"Error: Found table data on line {i+1} before separator line")
                return headers, data_rows, False
        
        elif table_found and separator_found:
            print(f"  Table ended after line {i}")
            break
    
    # Validate table structure
    if not header_found:
        print("Error: No Markdown table header found")
        return headers, data_rows, False
    
    if not separator_found:
        print("Error: No separator line found after header")
        return headers, data_rows, False
    
    if not data_rows:
        print("Warning: No data rows found in table")
    
    return headers, data_rows, True


def create_formatted_excel(df: pd.DataFrame, excel_path: str) -> bool:
    """
    Create a formatted Excel file from a pandas DataFrame.
    
    Applies professional formatting including:
    - Bold headers
    - Text wrapping for all cells
    - Markdown formatting preservation (bold/italic)
    - Automatic column width adjustment
    - Top-left alignment for readability
    
    Args:
        df: The pandas DataFrame to export
        excel_path: Path where the Excel file should be saved
        
    Returns:
        True if successful, False if an error occurred
    """
    if not OPENPYXL_AVAILABLE:
        print("Error: openpyxl library is required for formatting")
        print("Install with: pip install openpyxl")
        return False
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Table Data"
        
        # Write headers and data
        ws.append(df.columns.tolist())
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)
        
        print("Applying Excel formatting...")
        
        # Define styles
        header_font = Font(bold=True)
        wrap_alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        # Format header row
        for cell in ws[1]:
            if cell.value is not None:
                cleaned_value, _ = apply_cell_formatting(cell, str(cell.value))
                cell.value = cleaned_value
                cell.font = header_font
            cell.alignment = wrap_alignment
        
        # Format data rows
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if cell.value is not None:
                    cleaned_value, font_style = apply_cell_formatting(cell, str(cell.value))
                    cell.value = cleaned_value
                    if font_style:
                        cell.font = font_style
                
                cell.alignment = wrap_alignment
        
        # Adjust column widths
        print("Adjusting column widths...")
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                if cell.value:
                    try:
                        cell_lines = str(cell.value).split('\n')
                        max_line_length = max(len(line) for line in cell_lines) if cell_lines else 0
                        max_length = max(max_length, max_line_length)
                    except Exception:
                        continue
            
            # Set reasonable column width (max 60 characters, min 8)
            adjusted_width = min((max_length + 3) * 1.1, 60)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 8)
        
        # Ensure output directory exists
        output_dir = os.path.dirname(excel_path)
        if output_dir and not os.path.exists(output_dir):
            print(f"Creating output directory: {output_dir}")
            os.makedirs(output_dir)
        
        # Save the workbook
        wb.save(excel_path)
        print(f"Successfully saved formatted Excel file: '{excel_path}'")
        return True
        
    except Exception as e:
        print(f"Error creating Excel file '{excel_path}': {e}")
        return False


def markdown_table_to_excel_with_formatting(markdown_path: str, excel_path: str) -> bool:
    """
    Convert the first Markdown table in a file to a formatted Excel spreadsheet.
    
    This is the main processing function that orchestrates the entire conversion
    process from reading the Markdown file to saving the formatted Excel output.
    
    Args:
        markdown_path: Path to the input Markdown file
        excel_path: Path for the output Excel file (.xlsx)
        
    Returns:
        True if conversion was successful, False otherwise
    """
    print(f"Reading Markdown file: {markdown_path}")
    
    try:
        with open(markdown_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
    except FileNotFoundError:
        print(f"Error: Input file not found: '{markdown_path}'")
        return False
    except Exception as e:
        print(f"Error reading file '{markdown_path}': {e}")
        return False
    
    # Parse the Markdown table
    headers, data_rows, success = parse_markdown_table(lines)
    
    if not success:
        return False
    
    # Create DataFrame
    try:
        df = pd.DataFrame(data_rows, columns=headers)
        print(f"Created DataFrame: {len(df)} rows × {len(df.columns)} columns")
    except Exception as e:
        print(f"Error creating DataFrame: {e}")
        return False
    
    # Create formatted Excel file
    return create_formatted_excel(df, excel_path)


def main() -> None:
    """
    Main entry point for the command-line interface.
    
    Parses command-line arguments and executes the Markdown to Excel conversion
    with appropriate error handling and user feedback.
    """
    parser = argparse.ArgumentParser(
        description='Convert the first Markdown table in a file to a formatted Excel (.xlsx) file.',
        epilog='Example: python md_to_excel.py -i table.md -o output.xlsx',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument(
        '-i', '--inputfile',
        required=True,
        help='Path to the input Markdown file',
        metavar='MARKDOWN_FILE'
    )
    
    parser.add_argument(
        '-o', '--outputfile',
        required=True,
        help='Path for the output Excel file (must end with .xlsx)',
        metavar='EXCEL_FILE'
    )
    
    args = parser.parse_args()
    
    # Validate output file extension
    if not args.outputfile.lower().endswith('.xlsx'):
        parser.error("Output file must have .xlsx extension")
    
    # Perform the conversion
    success = markdown_table_to_excel_with_formatting(args.inputfile, args.outputfile)
    
    if not success:
        exit(1)


if __name__ == "__main__":
    main()